import json
from openpyxl import load_workbook

# ── 1. CE region reference ──
with open("/home/claude/Interreg-dashboard/src/data/data.js", "r") as f:
    content = f.read()
json_str = content.replace("export const DATA = ", "").rstrip().rstrip(";")
poc_data = json.loads(json_str)

CE_CODES = set(r["id"] for r in poc_data["regions"])
CE_MAP = {r["id"]: {"name": r["name"], "country": r["country"], "countryName": r["countryName"]} for r in poc_data["regions"]}

print(f"CE NUTS2 codes: {len(CE_CODES)}")

# ── 2. Load KB (NOT read_only for better performance on targeted reads) ──
print("Loading KB Excel...")
wb = load_workbook("/mnt/user-data/uploads/TNCOOP_knowledge_database.xlsx", data_only=True)
ws_meta = wb['metadatabase']

# ── 3. Extract metadata ──
indicators_meta = {}
for col in range(2, ws_meta.max_column + 1):
    code_raw = str(ws_meta.cell(row=2, column=col).value or "")
    primary_code = code_raw.split(",")[0].strip()
    indicators_meta[primary_code] = {
        "code": code_raw,
        "name": str(ws_meta.cell(row=1, column=col).value or ""),
        "description": str(ws_meta.cell(row=5, column=col).value or ""),
        "type": str(ws_meta.cell(row=6, column=col).value or "").lower(),
        "temporalStart": str(ws_meta.cell(row=7, column=col).value or ""),
        "temporalEnd": str(ws_meta.cell(row=8, column=col).value or ""),
        "spatialLevel": str(ws_meta.cell(row=18, column=col).value or ""),
        "source": str(ws_meta.cell(row=19, column=col).value or ""),
        "sourceUrl": str(ws_meta.cell(row=20, column=col).value or ""),
        "nutsVersion": str(ws_meta.cell(row=10, column=col).value or ""),
    }
print(f"Metadata: {len(indicators_meta)} indicators")

# ── 4. Theme mapping ──
THEME_MAPPING = [
    {"id": "demography-geography", "name": "Demography and geography", "icon": "👥",
     "subThemes": [
         {"id": "population-density", "name": "Population density", "type": "tncoop", "sheets": ["2a_Popdens"], "metaKey": "2a_Popdens"},
         {"id": "median-age", "name": "Median population age", "type": "tncoop", "sheets": ["3a_MedAge"], "metaKey": "3a_MedAge"},
         {"id": "population-change", "name": "Population change", "type": "tncoop", "sheets": ["12a_PopCha"], "metaKey": "12a_PopCha"},
    ]},
    {"id": "innovation-research-smes", "name": "Innovation, research and SMEs", "icon": "💡",
     "subThemes": [
         {"id": "competitiveness", "name": "Regional competitiveness", "type": "tncoop", "sheets": ["23a_RCI"], "metaKey": "23a_RCI"},
         {"id": "innovation", "name": "Regional innovation performance", "type": "tncoop", "sheets": ["25a_RIS"], "metaKey": "25a_RIS"},
    ]},
    {"id": "environment-climate", "name": "Environment and climate change", "icon": "🌍",
     "subThemes": [
         {"id": "waste", "name": "Waste intensity", "type": "tncoop", "sheets": ["153a_Waste"], "metaKey": "153a_Waste"},
         {"id": "energy", "name": "Renewable energy share", "type": "tncoop", "sheets": ["47a_Renewable"], "metaKey": "47a_Renewable"},
         {"id": "drought", "name": "Drought impact on ecosystems", "type": "tncoop", "sheets": ["155a_Drought"], "metaKey": "155a_Drought"},
         {"id": "flood", "name": "Flood risk", "type": "tncoop", "sheets": ["67a_Flood"], "metaKey": "67a_Flood"},
         {"id": "air-quality", "name": "Air quality (PM2.5)", "type": "tncoop", "sheets": ["52a_AirQ"], "metaKey": "52a_AirQ"},
         {"id": "natura-2000", "name": "Natura 2000 areas", "type": "tncoop", "sheets": [], "metaKey": None},
         {"id": "ghgs", "name": "Greenhouse gas emissions", "type": "tncoop", "sheets": [], "metaKey": None},
         {"id": "land-sealing", "name": "Land sealing", "type": "tncoop", "sheets": [], "metaKey": None},
    ]},
    {"id": "digital-connectivity-transport", "name": "Digital connectivity and transport", "icon": "🌐",
     "subThemes": [
         {"id": "broadband", "name": "Broadband access", "type": "tncoop", "sheets": ["86a_Broadband"], "metaKey": "86a_Broadband"},
         {"id": "road-accessibility", "name": "Road network accessibility", "type": "tncoop", "sheets": ["97a_TENT"], "metaKey": "97a_TENT"},
    ]},
    {"id": "sustainable-regional-development", "name": "Sustainable regional development", "icon": "🏗️",
     "subThemes": [
         {"id": "employment-sectors", "name": "Employment by economic sector", "type": "tncoop", "sheets": ["99g_EmplSect", "99h_EmplSect", "99i_EmplSect"], "metaKey": "99g_EmplSect"},
         {"id": "gdp", "name": "GDP per capita", "type": "tncoop", "sheets": ["156a_GDP"], "metaKey": "156a_GDP"},
         {"id": "governance", "name": "Quality of governance", "type": "tncoop", "sheets": ["154a_QualityGov"], "metaKey": "154a_QualityGov"},
    ]},
    {"id": "cultural-heritage-tourism", "name": "Cultural heritage and tourism", "icon": "🏛️",
     "subThemes": [
         {"id": "tourism", "name": "Tourism intensity", "type": "tncoop", "sheets": ["116a_Tourism"], "metaKey": "116a_Tourism"},
         {"id": "cultural-assets", "name": "Cultural assets density", "type": "tncoop", "sheets": ["157a_UNESCO"], "metaKey": "157a_UNESCO"},
    ]},
    {"id": "housing", "name": "Housing", "icon": "🏠",
     "subThemes": [
         {"id": "housing-prices", "name": "Housing prices", "type": "tncoop", "sheets": ["121a_Housing", "121b_Housing"], "metaKey": "121a_Housing"},
         {"id": "spatial-accessibility", "name": "Spatial accessibility to services", "type": "tncoop", "sheets": [], "metaKey": "134a_SpAccess"},
    ]},
    {"id": "people-to-people", "name": "People to people action and engagement", "icon": "🤝",
     "subThemes": [
         {"id": "education", "name": "Education attainment", "type": "tncoop", "sheets": ["101a_Edu", "101b_Edu", "101c_Edu"], "metaKey": "101a_Edu"},
         {"id": "neet", "name": "Youth not in employment or training", "type": "tncoop", "sheets": ["106a_NEET"], "metaKey": "106a_NEET"},
         {"id": "unemployment", "name": "Unemployment rate", "type": "tncoop", "sheets": ["107a_Unempl", "107b_Unempl", "107c_Unempl"], "metaKey": "107a_Unempl"},
         {"id": "poverty", "name": "People at risk of poverty", "type": "tncoop", "sheets": ["137a_AROPE"], "metaKey": "137a_AROPE"},
    ]},
]

SUB_LABELS = {
    "99g_EmplSect": "Primary sector", "99h_EmplSect": "Secondary sector", "99i_EmplSect": "Tertiary sector",
    "101a_Edu": "Primary & lower secondary (ISCED 0-2)", "101b_Edu": "Upper secondary & post-secondary (ISCED 3-4)", "101c_Edu": "Tertiary (ISCED 5-8)",
    "107a_Unempl": "Total", "107b_Unempl": "Male", "107c_Unempl": "Female",
    "121a_Housing": "Buying price (€/sqm)", "121b_Housing": "Renting price (€/sqm)",
}

# ── 5. Collect needed sheets ──
needed_sheets = set()
for t in THEME_MAPPING:
    for s in t["subThemes"]:
        needed_sheets.update(s["sheets"])

print(f"Sheets to extract: {len(needed_sheets)}")

# ── 6. Extract data per sheet ──
all_data = {}
for sheet_name in sorted(needed_sheets):
    if sheet_name not in wb.sheetnames:
        print(f"  ⚠ {sheet_name}: MISSING")
        continue
    ws = wb[sheet_name]
    # Headers
    headers = []
    for c in range(5, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h:
            headers.append((c, str(h).replace("Y_", "")))
    # Rows
    regions = []
    for row_idx in range(2, ws.max_row + 1):
        code = ws.cell(row=row_idx, column=1).value
        if not code or code not in CE_CODES:
            continue
        series = {}
        for col_idx, year in headers:
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                try:
                    series[year] = round(float(v), 6)
                except:
                    pass
        ref = CE_MAP.get(code, {})
        regions.append({
            "code": code,
            "name": ref.get("name", ""),
            "country": ref.get("country", ""),
            "series": series,
        })
    years = sorted(set(y for r in regions for y in r["series"].keys()))
    all_data[sheet_name] = {"years": years, "regions": regions}
    print(f"  ✓ {sheet_name}: {len(regions)} regions, {len(years)} years")

wb.close()

# ── 7. Build catalogue JS ──
themes_json = []
for theme in THEME_MAPPING:
    t = {"id": theme["id"], "name": theme["name"], "icon": theme["icon"], "subThemes": []}
    for sub in theme["subThemes"]:
        meta = None
        if sub["metaKey"] and sub["metaKey"] in indicators_meta:
            m = indicators_meta[sub["metaKey"]]
            meta = {k: m[k] for k in ["code","description","type","spatialLevel","temporalStart","temporalEnd","source","sourceUrl","nutsVersion"]}
        sub_inds = []
        for sh in sub["sheets"]:
            sub_inds.append({"sheetCode": sh, "label": SUB_LABELS.get(sh, sub["name"]), "hasData": sh in all_data and len(all_data[sh]["regions"]) > 0})
        st = {"id": sub["id"], "name": sub["name"], "type": sub["type"], "dataSheets": sub["sheets"], "metadata": meta,
              "subIndicators": sub_inds if len(sub_inds) > 1 else [], "hasData": any(sh in all_data and len(all_data[sh]["regions"]) > 0 for sh in sub["sheets"])}
        t["subThemes"].append(st)
    themes_json.append(t)

out = "/mnt/user-data/outputs/interreg-us"
with open(f"{out}/indicators-catalogue.js", "w") as f:
    f.write("// Auto-generated from TNCOOP_knowledge_database.xlsx\n// DO NOT EDIT\n\n")
    f.write("export const THEMES = " + json.dumps(themes_json, indent=2, ensure_ascii=False) + ";\n\n")
    f.write(open("/home/claude/helpers_cat.js").read() if False else "")
    f.write('export function getAllThemes() { return THEMES; }\n')
    f.write('export function getThemeById(id) { return THEMES.find(t => t.id === id) || null; }\n')
    f.write('export function getSubThemeById(themeId, subId) { const t = getThemeById(themeId); return t ? t.subThemes.find(s => s.id === subId) || null : null; }\n')
    f.write('export function getSubThemesByType(themeId, type) { const t = getThemeById(themeId); return t ? t.subThemes.filter(s => s.type === type) : []; }\n')

with open(f"{out}/indicators-data.js", "w") as f:
    f.write("// Auto-generated from TNCOOP_knowledge_database.xlsx\n// CE NUTS2 only (81 regions, 9 countries)\n// DO NOT EDIT\n\n")
    f.write("export const INDICATORS_DATA = " + json.dumps(all_data, indent=2, ensure_ascii=False) + ";\n\n")
    f.write('export function getIndicatorData(sheetCode) { return INDICATORS_DATA[sheetCode] || null; }\n')
    f.write('export function getIndicatorYears(sheetCode) { const d = INDICATORS_DATA[sheetCode]; return d ? d.years : []; }\n')

import os
c_size = os.path.getsize(f"{out}/indicators-catalogue.js") / 1024
d_size = os.path.getsize(f"{out}/indicators-data.js") / 1024
print(f"\n✅ Done!")
print(f"  indicators-catalogue.js: {c_size:.0f} KB")
print(f"  indicators-data.js: {d_size:.0f} KB")
print(f"  Themes: {len(themes_json)}, Sub-themes: {sum(len(t['subThemes']) for t in themes_json)}")
print(f"  With data: {sum(1 for t in themes_json for s in t['subThemes'] if s['hasData'])}")

